"""
File Downloader — Save attachments and download files from URLs.

Handles:
- Saving Gmail attachments to disk
- Downloading bảng kê PDFs from vinvoice.viettel.vn
- Duplicate detection
- Retry logic with exponential backoff

Skills applied: 03_web-scraper, 05_async-python-patterns, 09_error-handling-patterns
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from urllib.parse import unquote, urlparse

import requests
from tenacity import (
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential,
)

from src.models import DownloadResult, DownloadStatus, FileError, NetworkError

logger = logging.getLogger("email_auto_download.file_downloader")

REQUEST_TIMEOUT = 30  # seconds
MAX_RETRY_ATTEMPTS = 3
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)


class FileDownloader:
    """
    Download and save files from various sources.

    Features:
    - Duplicate detection by filename
    - Retry on network errors
    - Organized folder structure by subfolder
    """

    def __init__(
        self,
        output_dir: Path,
        skip_duplicates: bool = True,
        max_retries: int = MAX_RETRY_ATTEMPTS,
    ):
        self.output_dir = Path(output_dir)
        self.skip_duplicates = skip_duplicates
        self.max_retries = max_retries
        self._session: requests.Session | None = None

    def save_attachment(
        self,
        data: bytes,
        filename: str,
        subfolder: str = "",
    ) -> DownloadResult:
        """
        Save email attachment data to disk.

        Args:
            data: Raw file bytes
            filename: Target filename
            subfolder: Optional subfolder within output_dir

        Returns:
            DownloadResult with status and file path
        """
        if not data:
            return DownloadResult(
                status=DownloadStatus.FAILED,
                filepath=None,
                filename=filename,
                error_message="Empty attachment data",
            )

        # Check duplicate
        if self.skip_duplicates and self.is_duplicate(filename, subfolder):
            logger.info(f"⏭ Skipped duplicate: {filename}")
            return DownloadResult(
                status=DownloadStatus.SKIPPED_DUPLICATE,
                filepath=self._get_filepath(filename, subfolder),
                filename=filename,
                size_bytes=len(data),
            )

        try:
            filepath = self._get_filepath(filename, subfolder)
            filepath.parent.mkdir(parents=True, exist_ok=True)
            filepath.write_bytes(data)

            size_kb = len(data) / 1024
            logger.info(f"✅ Saved: {filename} ({size_kb:.0f} KB)")

            return DownloadResult(
                status=DownloadStatus.SUCCESS,
                filepath=filepath,
                filename=filename,
                size_bytes=len(data),
            )
        except OSError as e:
            logger.error(f"Failed to save {filename}: {e}")
            return DownloadResult(
                status=DownloadStatus.FAILED,
                filepath=None,
                filename=filename,
                error_message=str(e),
            )

    # Allowed domains for URL downloads (SSRF protection)
    _ALLOWED_DOMAINS: list[str] = [
        "vinvoice.viettel.vn",
        "s1.viettelpost.vn",
        "viettelpost.vn",
    ]

    def _is_url_allowed(self, url: str) -> bool:
        """Validate URL against allowed domain list to prevent SSRF."""
        try:
            parsed = urlparse(url)
            hostname = parsed.hostname or ""
            scheme = parsed.scheme.lower()
            if scheme not in ("http", "https"):
                return False
            return any(hostname.endswith(d) for d in self._ALLOWED_DOMAINS)
        except Exception:
            return False

    def download_from_url(
        self,
        url: str,
        filename: str | None = None,
        subfolder: str = "",
    ) -> DownloadResult:
        """
        Download a file from URL with SSRF protection.

        Only allows downloads from trusted domains.

        Args:
            url: Full URL to download
            filename: Optional target filename (auto-detect if None)
            subfolder: Optional subfolder within output_dir

        Returns:
            DownloadResult with status and file path
        """
        if not self._is_url_allowed(url):
            logger.warning(f"URL blocked by domain filter: {url}")
            return DownloadResult(
                status=DownloadStatus.FAILED,
                filepath=None,
                filename=filename or "unknown",
                error_message=f"Domain not in allowed list. URL: {url[:80]}",
            )

        try:
            response = self._http_get(url)

            # Domains known to return Excel directly — skip HTML guard
            _trusted_excel_domains = (
                "s1.viettelpost.vn",
                "viettelpost.vn",
            )
            skip_html_guard = any(d in url for d in _trusted_excel_domains)

            # ── Guard: Detect HTML response (login page / error page) ──
            if not skip_html_guard:
                ct = response.headers.get("Content-Type", "").lower()
                content_preview = response.content[:200]
                is_html = (
                    "text/html" in ct
                    or content_preview.lstrip().startswith(b"<!DOCTYPE")
                    or content_preview.lstrip().startswith(b"<html")
                )
                if is_html:
                    logger.warning(
                        "Bảng kê URL returned HTML (likely requires browser auth). "
                        "URL: %s", url
                    )
                    return DownloadResult(
                        status=DownloadStatus.FAILED,
                        filepath=None,
                        filename=filename or "bangke",
                        error_message=(
                            "Server trả về trang HTML (cần đăng nhập). "
                            "Viettel Portal yêu cầu session/cookie từ trình duyệt."
                        ),
                    )

            # Detect filename
            if not filename:
                filename = self._extract_filename(response, url)

            # For trusted Viettel domains: convert HTML bảng kê → real xlsx
            if skip_html_guard:
                ct = response.headers.get("Content-Type", "").lower()
                is_html = (
                    "text/html" in ct
                    or response.content[:100].lstrip().startswith(b"<")
                )
                if is_html:
                    # Make sure filename ends with .xlsx
                    stem = Path(filename).stem
                    unique = self._extract_url_id(url)
                    xlsx_name = f"{stem}_{unique}.xlsx" if unique else f"{stem}.xlsx"

                    if self.skip_duplicates and self.is_duplicate(xlsx_name, subfolder):
                        logger.info(f"⏭ Skipped duplicate: {xlsx_name}")
                        return DownloadResult(
                            status=DownloadStatus.SKIPPED_DUPLICATE,
                            filepath=self._get_filepath(xlsx_name, subfolder),
                            filename=xlsx_name,
                            size_bytes=0,
                        )

                    filepath = self._get_filepath(xlsx_name, subfolder)
                    filepath.parent.mkdir(parents=True, exist_ok=True)
                    ok = self._html_to_xlsx(response.content.decode("utf-8", errors="replace"), filepath)
                    if ok:
                        logger.info(f"✅ Converted HTML → xlsx: {xlsx_name}")
                        return DownloadResult(
                            status=DownloadStatus.SUCCESS,
                            filepath=filepath,
                            filename=xlsx_name,
                            size_bytes=filepath.stat().st_size,
                        )
                    else:
                        return DownloadResult(
                            status=DownloadStatus.FAILED,
                            filepath=None,
                            filename=xlsx_name,
                            error_message="HTML→xlsx conversion failed",
                        )

            # Check duplicate
            if self.skip_duplicates and self.is_duplicate(filename, subfolder):
                logger.info(f"⏭ Skipped duplicate: {filename}")
                return DownloadResult(
                    status=DownloadStatus.SKIPPED_DUPLICATE,
                    filepath=self._get_filepath(filename, subfolder),
                    filename=filename,
                    size_bytes=len(response.content),
                )

            # Save file as-is
            return self.save_attachment(response.content, filename, subfolder)

        except (NetworkError, requests.RequestException) as e:
            logger.error(f"Failed to download from URL: {e}")
            return DownloadResult(
                status=DownloadStatus.FAILED_RETRY_EXHAUSTED,
                filepath=None,
                filename=filename or "unknown",
                error_message=str(e),
            )

    def is_duplicate(self, filename: str, subfolder: str = "") -> bool:
        """Check if file already exists on disk."""
        filepath = self._get_filepath(filename, subfolder)
        return filepath.exists()

    def _html_to_xlsx(self, html: str, output_path: Path) -> bool:
        """Parse Viettel bảng kê HTML and write a real Excel file.

        Extracts:
        - Header info (invoice number, company, date range)
        - Detail table (#tdetail) with all shipment rows

        Returns True on success.
        """
        try:
            from bs4 import BeautifulSoup
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            soup = BeautifulSoup(html, "lxml")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bảng kê chi tiết"

            current_row = 1

            # ── Header info from page ─────────────────────────────────
            header_style = Font(bold=True, size=11)
            title_fill = PatternFill("solid", fgColor="FFA500")  # orange

            def _text(selector_fn) -> str:
                el = selector_fn()
                return el.get_text(strip=True) if el else ""

            # Title
            title_el = soup.find("span", style=lambda s: s and "font-weight: bold" in s and "BẢNG KÊ" in (soup.find("span", style=lambda s2: s2 and "font-weight: bold" in s2) or soup).get_text())
            # Simpler: find spans containing known text
            for span in soup.find_all("span"):
                text = span.get_text(strip=True)
                if text in ("BẢNG KÊ CHI TIẾT",):
                    cell = ws.cell(row=current_row, column=1, value=text)
                    cell.font = Font(bold=True, size=14)
                    current_row += 1
                    break

            # Extract key info spans from the top table (before #tdetail)
            top_table = soup.find("table", style=lambda s: s and "border: 0" in s)
            if top_table:
                for span in top_table.find_all("span"):
                    text = span.get_text(strip=True)
                    if text and len(text) > 2 and "BẢNG KÊ" not in text:
                        cell = ws.cell(row=current_row, column=1, value=text)
                        cell.font = Font(size=10)
                        current_row += 1

            current_row += 1  # blank row

            # ── Detail table ──────────────────────────────────────────
            detail_table = soup.find("table", id="tdetail")
            if detail_table:
                # Header row
                thead = detail_table.find("thead")
                if thead:
                    headers = [td.get_text(strip=True) for td in thead.find_all("td")]
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=h)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="1F497D")
                        cell.alignment = Alignment(horizontal="center", wrap_text=True)
                    current_row += 1

                # Data rows
                tbody = detail_table.find("tbody")
                if tbody:
                    # Columns that should stay as text (not parsed as numbers)
                    _text_cols = {"stt", "phiếu gửi", "biển kiểm soát",
                                  "nơi đi", "nơi đến", "dịch vụ", "ngày gửi"}
                    text_col_indices = set()
                    if thead:
                        for ci, td in enumerate(thead.find_all("td"), 1):
                            if td.get_text(strip=True).lower() in _text_cols:
                                text_col_indices.add(ci)

                    from openpyxl.utils import get_column_letter
                    for tr in tbody.find_all("tr"):
                        cols = tr.find_all("td")
                        for col, td in enumerate(cols, 1):
                            val = td.get_text(strip=True)
                            cell = ws.cell(row=current_row, column=col)
                            if col in text_col_indices or not val:
                                # Force text format
                                cell.value = val
                                cell.number_format = "@"
                            else:
                                # Try to parse numbers (cước phí, tiền thuế...)
                                num_val = val.replace(",", "")
                                try:
                                    cell.value = int(num_val) if num_val.isdigit() else val
                                except (ValueError, AttributeError):
                                    cell.value = val
                        current_row += 1

                # Footer/totals row
                tfoot = detail_table.find("tfoot")
                if tfoot:
                    current_row += 1
                    for tr in tfoot.find_all("tr"):
                        cols = tr.find_all("td")
                        for col, td in enumerate(cols, 1):
                            val = td.get_text(strip=True)
                            if val:
                                cell = ws.cell(row=current_row, column=col, value=val)
                                cell.font = Font(bold=True)
                    current_row += 1

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            wb.save(output_path)
            return True

        except Exception as e:
            logger.error(f"_html_to_xlsx failed: {e}", exc_info=True)
            return False


    def get_download_count(self, subfolder: str = "") -> int:
        """Count files in a subfolder."""
        target_dir = self.output_dir / subfolder if subfolder else self.output_dir
        if not target_dir.exists():
            return 0
        return sum(1 for f in target_dir.iterdir() if f.is_file())

    # ── Private helpers ──────────────────────────────────────────────

    def _get_filepath(self, filename: str, subfolder: str = "") -> Path:
        """Build full file path from components."""
        # Sanitize filename
        filename = self._sanitize_filename(filename)
        if subfolder:
            return self.output_dir / subfolder / filename
        return self.output_dir / filename

    @retry(
        stop=stop_after_attempt(MAX_RETRY_ATTEMPTS),
        wait=wait_exponential(multiplier=2, min=3, max=30),
        retry=retry_if_exception_type((ConnectionError, TimeoutError)),
        reraise=True,
    )
    def _http_get(self, url: str) -> requests.Response:
        """HTTP GET with retry logic."""
        session = self._get_session()
        response = session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        response.raise_for_status()
        return response

    def _get_session(self) -> requests.Session:
        """Get or create an HTTP session with proper headers."""
        if self._session is None:
            self._session = requests.Session()
            self._session.headers.update({
                "User-Agent": USER_AGENT,
                "Accept": "*/*",
            })
        return self._session

    # Map MIME types → file extensions
    _MIME_TO_EXT: dict[str, str] = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/vnd.ms-excel": ".xlsx",
        "application/vnd.ms-excel.sheet.macroenabled.12": ".xlsm",
        "application/pdf": ".pdf",
        "application/xml": ".xml",
        "text/xml": ".xml",
        "application/zip": ".zip",
        "text/csv": ".csv",
    }

    def _extract_filename(self, response: requests.Response, url: str) -> str:
        """Extract filename from response headers or URL.

        Priority:
        1. Content-Disposition header  (most reliable)
        2. Detect real type via Content-Type and fix extension
        3. URL path basename
        """
        # 1. Content-Disposition header
        cd = response.headers.get("Content-Disposition", "")
        if cd:
            match = re.findall(r'filename[*]?=["\']?([^"\';\r\n]+)', cd)
            if match:
                name = unquote(match[0].strip())
                # Even if header gives a filename, fix .do extension
                return self._fix_extension(name, response, url)

        # 2. URL path basename
        parsed = urlparse(url)
        path_filename = unquote(Path(parsed.path).name)

        if path_filename:
            return self._fix_extension(path_filename, response, url)

        # 3. Last resort: generate name from content type
        ext = self._ext_from_content_type(response, url)
        return f"bangke_chi_tiet{ext}"

    def _fix_extension(self, filename: str, response: requests.Response, original_url: str = "") -> str:
        """Replace bad extensions (e.g. .do) with the real one from Content-Type.
        Also appends unique suffix from URL params (invoiceId) to avoid filename collisions.
        """
        stem = Path(filename).stem
        current_ext = Path(filename).suffix.lower()

        # Extensions that are actually server endpoints, not real file types
        _fake_exts = {".do", ".action", ".aspx", ".jsp", ".php", ""}

        if current_ext in _fake_exts:
            real_ext = self._ext_from_content_type(response, original_url)
            if real_ext:
                logger.debug(
                    f"Replacing fake extension '{current_ext}' → '{real_ext}' "
                    f"for '{filename}'"
                )
                # Add a unique suffix from invoiceId query param if present
                unique = self._extract_url_id(original_url)
                name = f"{stem}_{unique}{real_ext}" if unique else f"{stem}{real_ext}"
                return name

        return filename

    @staticmethod
    def _extract_url_id(url: str) -> str:
        """Extract a short unique ID from URL query params (invoiceId or similar)."""
        from urllib.parse import urlparse, parse_qs
        try:
            params = parse_qs(urlparse(url).query)
            for key in ("invoiceId", "id", "invoice_id"):
                if key in params:
                    raw = params[key][0]
                    # Use last 8 chars of the param as a short unique suffix
                    safe = re.sub(r"[^A-Za-z0-9_-]", "", raw)
                    return safe[-12:] if len(safe) > 12 else safe
        except Exception:
            pass
        return ""

    def _ext_from_content_type(self, response: requests.Response, url: str = "") -> str:
        """Map Content-Type header to a file extension.
        For known Viettel download domains, default to .xlsx when type is unknown.
        """
        ct = response.headers.get("Content-Type", "").split(";")[0].strip().lower()
        if ct in self._MIME_TO_EXT:
            return self._MIME_TO_EXT[ct]
        # Trusted Viettel Excel endpoints — use .xlsx as default
        _viettel_excel_domains = ("s1.viettelpost.vn", "viettelpost.vn")
        if any(d in url for d in _viettel_excel_domains):
            return ".xlsx"
        return ".bin"  # Unknown type — safe fallback



    def _sanitize_filename(self, filename: str) -> str:
        """Remove unsafe characters and path components from filename."""
        # Strip path components to prevent path traversal (../../etc)
        filename = Path(filename).name
        # Remove/replace characters not safe for Windows filenames
        filename = re.sub(r'[<>:"/\\|?*]', "_", filename)
        filename = filename.strip(". ")
        return filename or "unnamed_file"

    def close(self) -> None:
        """Close HTTP session."""
        if self._session:
            self._session.close()
            self._session = None
